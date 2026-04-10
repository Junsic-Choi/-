import openpyxl

file_path = r"c:\Users\i0215099\Desktop\ANTI_TEST\정삭 장비 우선 순위(HSP HM2J AH2J 10T 15T).xlsx"
print(f"Loading {file_path}")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print("Row 1 (Header):")
    for cell in ws[1][:10]:
        fill = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else "None"
        font = cell.font.color.rgb if cell.font and cell.font.color else "None"
        bold = cell.font.bold if cell.font else "None"
        h_align = cell.alignment.horizontal if cell.alignment else "None"
        v_align = cell.alignment.vertical if cell.alignment else "None"
        borders = []
        if cell.border:
            if cell.border.top.style: borders.append("top")
            if cell.border.bottom.style: borders.append("bottom")
            if cell.border.left.style: borders.append("left")
            if cell.border.right.style: borders.append("right")
        print(f"  {cell.coordinate} ({cell.value}): fill={fill}, font_color={font}, bold={bold}, align={h_align}/{v_align}, borders={'-'.join(borders)}")

    print("\nRow 2 (Data):")
    for cell in ws[2][:10]:
        fill = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else "None"
        font = cell.font.color.rgb if cell.font and cell.font.color else "None"
        bold = cell.font.bold if cell.font else "None"
        h_align = cell.alignment.horizontal if cell.alignment else "None"
        v_align = cell.alignment.vertical if cell.alignment else "None"
        borders = []
        if cell.border:
            if cell.border.top.style: borders.append("top")
            if cell.border.bottom.style: borders.append("bottom")
            if cell.border.left.style: borders.append("left")
            if cell.border.right.style: borders.append("right")
        print(f"  {cell.coordinate} ({cell.value}): fill={fill}, font_color={font}, bold={bold}, align={h_align}/{v_align}, borders={'-'.join(borders)}")

    print("\nRow 3 (Data):")
    for cell in ws[3][:10]:
        fill = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else "None"
        font = cell.font.color.rgb if cell.font and cell.font.color else "None"
        bold = cell.font.bold if cell.font else "None"
        h_align = cell.alignment.horizontal if cell.alignment else "None"
        v_align = cell.alignment.vertical if cell.alignment else "None"
        borders = []
        if cell.border:
            if cell.border.top.style: borders.append("top")
            if cell.border.bottom.style: borders.append("bottom")
            if cell.border.left.style: borders.append("left")
            if cell.border.right.style: borders.append("right")
        print(f"  {cell.coordinate} ({cell.value}): fill={fill}, font_color={font}, bold={bold}, align={h_align}/{v_align}, borders={'-'.join(borders)}")

except Exception as e:
    print(f"Error: {e}")
